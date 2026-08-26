"""
eval/human_eval_01_generate_template.py

Builds the v2 human-labeling kit for a manual reviewer (RA):
  1. labeling_template_v2.xlsx      -- BLIND (no model predictions visible).
                                        Give this to the RA. Dropdown-validated
                                        input columns; full observation text;
                                        an Instructions tab.
  2. DO_NOT_SHARE_answer_key_v2.csv -- model predictions, keyed by (fei, obs_num).
                                        Keep private. Used only by
                                        human_eval_02_score.py after labels come back.

Why blind: showing the RA the model's own prediction while they label invites
anchoring bias and inflates agreement metrics in a way a reviewer could
reasonably challenge. Predictions are joined back in only for scoring.

Field definitions match Yelena Ionova's expert-review revision (v2 prompt) --
see 483_Labeling_Rules_v2.docx, which is a plain-language rewrite of the same
rules given to the LLM (01_extract_observation_signals.py, --prompt-version v2),
so RA labels and LLM output are directly comparable.

Run from this folder:
  python human_eval_01_generate_template.py
  python human_eval_01_generate_template.py --source <path-to-a-different-v2-signals.csv>
  python human_eval_01_generate_template.py --sample 80 --stratify-severity
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
DATA = HERE.parent

# Default source: the Sonnet 5 v2 model-comparison sample (n=50, round-robin
# across FEIs). Real, already-scored data -- costs nothing further to reuse
# as the RA's first batch. Not severity-stratified; pass --stratify-severity
# with a larger --sample against a bigger v2 signals file for tier balance.
DEFAULT_SOURCE = DATA / "483_observation_context_signals_sample50_claudesonnet5_v2.csv"

TEMPLATE_XLSX  = HERE / "labeling_template_v2.xlsx"
ANSWER_KEY_CSV = HERE / "DO_NOT_SHARE_answer_key_v2.csv"

VIOLATION_CATEGORIES = ["QualitySystem", "ProductionSystem", "MaterialsSystem",
                         "FacilitiesEquipmentSystem", "LaboratoryControlsSystem",
                         "PackagingLabelingSystem", "Other"]
SEVERITY_TIERS = ["Critical", "Major", "Moderate", "Minor"]
SCOPES = ["SingleBatch", "MultipleProducts", "FacilityWide", "Unclear"]
ROOT_CAUSES = ["Capital", "Cultural", "Mixed", "Unclear"]
REMEDIATION = ["Strong", "Partial", "Weak", "None"]
BOOL_VALS = ["TRUE", "FALSE"]
CONFIDENCE_VALS = ["1", "2", "3", "4", "5"]

ANSWER_KEY_COLS = [
    "fei", "obs_num", "violation_category", "severity_tier", "scope",
    "root_cause_type", "remediation_signal", "data_integrity_flag_llm",
    "repeat_flag_llm", "patient_risk_flag_llm", "contamination_flag_llm",
    "contamination_risk_flag_llm", "investigation_flag_llm",
    "patient_risk_rationale", "model_name", "prompt_version",
]

# Read-only context columns, then human_* input columns (dropdown-validated
# where the field is categorical/boolean).
COLUMNS = [
    ("row_id",                        12, False),
    ("fei",                           12, False),
    ("insp_date",                     12, False),
    ("obs_num",                        9, False),
    ("cfr_codes",                     14, False),
    ("observation_text",              70, False),
    ("human_violation_category",      26, True),
    ("human_severity_tier",           16, True),
    ("human_scope",                   18, True),
    ("human_root_cause_type",         16, True),
    ("human_remediation_signal",      18, True),
    ("human_repeat_flag",             14, True),
    ("human_patient_risk_flag",       16, True),
    ("human_patient_risk_why",        40, True),
    ("human_contamination_flag",      18, True),
    ("human_contamination_risk_flag", 22, True),
    ("human_contamination_why",       40, True),
    ("human_investigation_flag",      16, True),
    ("human_data_integrity_flag",     18, True),
    ("human_confidence_1to5",         14, True),
    ("notes",                         40, True),
]


def _load_source(source: Path, sample: int | None, stratify_severity: bool) -> pd.DataFrame:
    df = pd.read_csv(source)
    df = df[df["extraction_status"].isin(["ok", "partial"])].reset_index(drop=True)
    print(f"Source rows (ok/partial): {len(df)}  from {source.name}")

    if sample and sample < len(df):
        if stratify_severity and "severity_tier" in df.columns:
            n_each = max(1, sample // df["severity_tier"].nunique())
            parts = [g.sample(n=min(n_each, len(g)), random_state=7)
                     for _, g in df.groupby("severity_tier")]
            picked = pd.concat(parts)
            remaining = sample - len(picked)
            if remaining > 0:
                leftover = df.drop(picked.index)
                picked = pd.concat([picked, leftover.sample(
                    n=min(remaining, len(leftover)), random_state=8)])
            df = picked.sample(frac=1, random_state=9).reset_index(drop=True)
        else:
            df = df.sample(n=sample, random_state=7).reset_index(drop=True)
    return df


def _write_answer_key(df: pd.DataFrame) -> None:
    cols = [c for c in ANSWER_KEY_COLS if c in df.columns]
    out = df[cols].copy()
    if "remediation_signal" in out.columns:
        # "None" is a legitimate category value here (= no remediation
        # mentioned), not a missing-data marker -- but pandas' default
        # read_csv NA-sniffing treats the literal text "None" as NaN on
        # every round trip through CSV. For rows that reached this point
        # (extraction_status ok/partial), _validate() in
        # 01_extract_observation_signals.py guarantees remediation_signal is
        # always one of Strong/Partial/Weak/None -- so any NaN we see here is
        # unambiguously a mangled "None", never a genuinely missing value.
        # Same fix already applied in 02_aggregate_fei_features.py.
        out["remediation_signal"] = out["remediation_signal"].fillna("None")
    out.to_csv(ANSWER_KEY_CSV, index=False)
    print(f"Answer key written (PRIVATE -- do not share): {ANSWER_KEY_CSV}")


def _write_template(df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Labeling"

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    input_fill  = PatternFill("solid", fgColor="FFF9E6")  # pale yellow = fill in
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (name, width, _) in enumerate(COLUMNS, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, size=10)
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = border
    ws.freeze_panes = "G2"
    ws.row_dimensions[1].height = 28

    for r, (_, row) in enumerate(df.iterrows(), start=2):
        values = {
            "row_id": r - 1,
            "fei": row["fei"],
            "insp_date": row.get("insp_date", ""),
            "obs_num": row["obs_num"],
            "cfr_codes": row.get("cfr_codes", ""),
            "observation_text": str(row["obs_text_clean"]).strip(),
        }
        for i, (name, _, is_input) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=r, column=i, value=values.get(name, ""))
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if is_input:
                c.fill = input_fill
        ws.row_dimensions[r].height = 90

    last_row = len(df) + 1

    def add_dropdown(col_name, values):
        col_idx = [i for i, (n, _, _) in enumerate(COLUMNS, start=1) if n == col_name][0]
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=f'"{",".join(values)}"',
                             allow_blank=True, showDropDown=False)
        dv.error = "Please choose a value from the dropdown list."
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    add_dropdown("human_violation_category", VIOLATION_CATEGORIES)
    add_dropdown("human_severity_tier", SEVERITY_TIERS)
    add_dropdown("human_scope", SCOPES)
    add_dropdown("human_root_cause_type", ROOT_CAUSES)
    add_dropdown("human_remediation_signal", REMEDIATION)
    add_dropdown("human_repeat_flag", BOOL_VALS)
    add_dropdown("human_patient_risk_flag", BOOL_VALS)
    add_dropdown("human_contamination_flag", BOOL_VALS)
    add_dropdown("human_contamination_risk_flag", BOOL_VALS)
    add_dropdown("human_investigation_flag", BOOL_VALS)
    add_dropdown("human_data_integrity_flag", BOOL_VALS)
    add_dropdown("human_confidence_1to5", CONFIDENCE_VALS)

    # Instructions tab
    ws2 = wb.create_sheet("Instructions", 0)
    ws2.column_dimensions["A"].width = 100
    instructions = [
        ("483 Observation Labeling -- Instructions", True, 14),
        ("", False, 11),
        ("What this is: real FDA Form 483 observations from our facility inspection "
         "dataset. Assign the same 11 labels a domain expert would, using ONLY the "
         "observation text -- do not look up the facility, drug, or any outside "
         "information.", False, 11),
        ("", False, 11),
        ("Full field definitions, valid values, and worked examples are in "
         "483_Labeling_Rules_v2.docx. Read that first -- these are the exact same "
         "definitions our LLM pipeline uses, so your labels are directly comparable to "
         "its output.", False, 11),
        ("", False, 11),
        ("How to fill this in:", True, 12),
        ("  1. Work in the 'Labeling' tab. Columns A-F are read-only context (do not "
         "edit).", False, 11),
        ("  2. Columns with a pale yellow fill are yours to complete. Most are dropdown "
         "lists -- click the cell and choose from the arrow.", False, 11),
        ("  3. human_patient_risk_why: a short sentence on WHY you set the patient-risk "
         "flag the way you did. Required whenever the flag is TRUE.", False, 11),
        ("  4. human_contamination_why: a short sentence on WHY you set "
         "human_contamination_flag and human_contamination_risk_flag the way you did -- "
         "this is the newest, subtlest distinction in the rules (confirmed event vs. "
         "control-risk gap), so we want to see your reasoning even when you're confident.",
         False, 11),
        ("  5. human_confidence_1to5: how confident you are in your own labels for that "
         "row (1 = guessing, 5 = certain).", False, 11),
        ("  6. notes: anything ambiguous, anything you'd flag for discussion, or where "
         "the text itself seems incomplete/redacted in a way that affects your answer.",
         False, 11),
        ("", False, 11),
        ("Please work independently and do not discuss individual rows with anyone "
         "until you're done -- this keeps the comparison clean. If a row is genuinely "
         "ambiguous even after reading the manual, label it your best judgment and flag "
         "it in 'notes' rather than skipping it.", False, 11),
        ("", False, 11),
        ("When you're done, save the file and send it back -- do not change the file "
         "name or sheet names, and do not add/remove/reorder columns.", False, 11),
    ]
    r = 1
    for text, bold, size in instructions:
        c = ws2.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 34 if len(text) > 90 else (24 if text else 8)
        r += 1

    wb.save(TEMPLATE_XLSX)
    print(f"Blind labeling template written: {TEMPLATE_XLSX}")
    print(f"  Rows: {len(df)}")


def main():
    parser = argparse.ArgumentParser(description="Build the v2 human-labeling kit.")
    parser.add_argument("--source", type=str, default=str(DEFAULT_SOURCE),
                         help="Path to a step01 v2 signals CSV to sample from.")
    parser.add_argument("--sample", type=int, default=None,
                         help="Subsample this many rows from --source (default: use all).")
    parser.add_argument("--stratify-severity", action="store_true",
                         help="Stratify the subsample by severity_tier (needs --sample).")
    args = parser.parse_args()

    source = Path(args.source)
    if not source.exists():
        raise FileNotFoundError(f"Source signals file not found: {source}")

    df = _load_source(source, args.sample, args.stratify_severity)
    _write_answer_key(df)
    _write_template(df)


if __name__ == "__main__":
    main()
